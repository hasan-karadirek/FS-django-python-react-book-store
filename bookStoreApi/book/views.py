from rest_framework import status, permissions
from rest_framework.response import Response
from rest_framework.views import APIView
from .models import Book, Category, Language
from .serializers import (
    BookSerializer,
    CategorySerializer,
    LanguageSerializer,
    UpdateBooksSerializer,
)
from rest_framework.parsers import MultiPartParser, FormParser
from core.custom_exceptions import CustomAPIException
from core.mixins import IsBookExist
from core.helpers import pagination
from django.db.models import Q
import openpyxl
from io import BytesIO
from django.http import FileResponse
from django.utils import timezone
from django.core.files.uploadedfile import InMemoryUploadedFile
from django.db import transaction


# class AddBookAPIView(APIView):
#     parser_classes = (MultiPartParser, FormParser)  # To handle file uploads
#     permission_classes = [permissions.IsAdminUser]

#     def post(self, request, *args, **kwargs):
#         serializer = BookSerializer(data=request.data)

#         if serializer.is_valid():
#             serializer.save()
#             return Response(serializer.data, status=status.HTTP_201_CREATED)
#         else:
#             raise CustomAPIException(str(serializer.errors), status=400)


# class UpdateBookAPIView(IsBookExist, APIView):
#     parser_classes = (MultiPartParser, FormParser)
#     permission_classes = [permissions.IsAdminUser]

#     def put(self, request, pk, *args, **kwargs):

#         serializer = BookSerializer(request.book, data=request.data)

#         if serializer.is_valid():
#             serializer.save()
#             return Response(serializer.data, status=status.HTTP_200_OK)
#         else:
#             raise CustomAPIException(str(serializer.errors), status=400)


# class DeleteBookAPIView(IsBookExist, APIView):
#     permission_classes = [permissions.IsAdminUser]

#     def delete(self, request, pk, *args, **kwargs):

#         request.book.delete()
#         return Response(status=status.HTTP_204_NO_CONTENT)


class GetBookAPIView(IsBookExist, APIView):
    def get(self, request, slug, *args, **kwargs):

        serializer = BookSerializer(request.book)
        response = {"data": serializer.data, "success": True}
        return Response(response, status=status.HTTP_200_OK)


class GetBooksAPIView(APIView):
    def get(self, request, *args, **kwargs):

        # Subquery to check for open sales for a book
        category_query = request.query_params.get("category", None)
        language_query = request.query_params.get("language", None)
        tag_query = request.query_params.get("tag", None)

        filter_criteria = {"status": "OPEN"}
        if tag_query:
            filter_criteria["tag__name"] = tag_query
        if category_query:
            filter_criteria["category__title"] = category_query
        if language_query:
            filter_criteria["language__name"] = language_query
        books = Book.objects.filter(**filter_criteria)

        search_query = request.query_params.get("search", None)
        if search_query:
            books = books.filter(
                (
                    Q(title__icontains=search_query)
                    | Q(author__name__icontains=search_query)
                    | Q(isbn__icontains=search_query)
                    | Q(publishing_house__name__icontains=search_query)
                    | Q(tags__name__icontains=search_query)
                )
            ).distinct()
        books = books.order_by("-id")
        page_number = request.query_params.get("page", 1)
        paginated_data = pagination(books, 20, page_number)
        page = paginated_data["page"]

        serializer = BookSerializer(page, many=True)
        paginated_data["page"] = serializer.data

        response = {"data": paginated_data, "success": True}
        return Response(response, status=status.HTTP_200_OK)


class GetAllCategoriesApiView(APIView):
    def get(self, request, *args, **kwargs):

        categories = Category.objects.all().distinct()

        serializer = CategorySerializer(categories, many=True)
        response = {"data": serializer.data, "success": True}
        return Response(response, status=status.HTTP_200_OK)


class GetAllLanguagesApiView(APIView):
    def get(self, request, *args, **kwargs):
        languages = Language.objects.all().distinct()
        serializer = LanguageSerializer(languages, many=True)
        response = {"data": serializer.data, "success": True}
        return Response(response, status=status.HTTP_200_OK)


class UpdateBooksAPIView(APIView):
    permission_classes = [permissions.IsAdminUser]
    parser_classes = (MultiPartParser, FormParser)

    def put(self, request, *args, **kwargs):
        serializer = UpdateBooksSerializer(data=request.data)
        if serializer.is_valid():
            file = serializer.validated_data["file"]

            if isinstance(file, InMemoryUploadedFile):
                workbook = openpyxl.load_workbook(file)
                sheet = workbook.active

                issued_books = []
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    id, status = row
                    try:
                        book = Book.objects.get(id=id)
                        book.status = status
                        book.save()
                    except Book.DoesNotExist:
                        issued_books.append(id)

                return Response(
                    {"success": True, "msg": f"invalid ID list: {issued_books}"}
                )
        
        raise CustomAPIException("Request body is not valid", 400)

class UpdateBooksStatusAPIView(APIView):
    permission_classes = [permissions.IsAdminUser]

    @transaction.atomic
    def put(self, request, *args, **kwargs):
        book_ids = request.data.get("ids").replace(" ","").split(",")
        books_to_update = []
        for book_id in book_ids:
            if book_id.isdigit() == False:
                raise CustomAPIException(f"Invalid book ID {book_id}", status=status.HTTP_400_BAD_REQUEST)
            try:
                book = Book.objects.get(env_no=int(book_id))
                books_to_update.append(book)
            except Book.DoesNotExist:
                raise CustomAPIException("Book not found", status=status.HTTP_404_NOT_FOUND)
        for book in books_to_update:
            book.status = "SOLD"
            book.save()
        return Response({"success": True, "msg": "Books updated successfully"}, status=status.HTTP_200_OK)

class ExportBooksAPIView(APIView):
    permission_classes = [permissions.IsAdminUser]

    def get(self, request, *args, **kwargs):
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        header = [
            "id",
            "isbn",
            "env_no",
            "title",
            "author",
            "language",
            "cover",
            "publishing_house",
            "year",
            "edition",
            "category",
            "condition",
            "condition_description",
            "tags",
            "price",
            "entry",
            "status",
            "page",
            "ant",
            "cost",
            "remain",
            "loc",
            "supplier",
        ]
        sheet.append(header)

        books = Book.objects.all()

        for book in books:
            row = [
                book.id,
                book.isbn,
                book.env_no,
                book.title,
                book.author.name if book.author else "",
                book.language.name if book.language else "",
                book.cover,
                book.publishing_house.name if book.publishing_house else "",
                book.year,
                book.edition,
                book.category.title if book.category else "",
                book.condition,
                book.condition_description,
                ", ".join([tag.name for tag in book.tags.all()]),
                book.price,
                book.entry,
                book.status,
                book.page,
                book.ant,
                book.cost,
                book.remain,
                book.loc,
                book.supplier,
            ]
            sheet.append(row)

        file_stream = BytesIO()
        workbook.save(file_stream)
        file_stream.seek(0)

        response = FileResponse(
            file_stream,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response[
            "Content-Disposition"
        ] = f'attachment; filename=books_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

        return response
